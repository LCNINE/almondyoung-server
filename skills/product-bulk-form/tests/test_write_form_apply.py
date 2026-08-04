import openpyxl
import pytest

from read_form import read_form
from write_form import apply_changes


def test_적지_않은_필드는_건드리지_않는다(prefilled_workbook, columns, tmp_path):
    out = tmp_path / "edited.xlsx"
    apply_changes(prefilled_workbook, {"변경": [{"상품키": "P-000001", "필드": {"brand": "NEW"}}]}, out, columns)

    product = read_form(out, columns)["products"][0]
    assert product["필드"]["brand"] == "NEW"
    assert product["필드"]["name"] == "티셔츠"      # 안 건드림
    assert product["필드"]["basePrice"] == "19000"  # 안 건드림


def test_null_은_명시적_비움이다(prefilled_workbook, columns, tmp_path):
    out = tmp_path / "edited.xlsx"
    apply_changes(prefilled_workbook, {"변경": [{"상품키": "P-000001", "필드": {"brand": None}}]}, out, columns)

    assert read_form(out, columns)["products"][0]["필드"]["brand"] == ""


def test_카테고리_행목록은_교체된다(prefilled_workbook, columns, tmp_path):
    out = tmp_path / "edited.xlsx"
    apply_changes(
        prefilled_workbook,
        {"변경": [{"상품키": "P-000001",
                   "카테고리": [{"categoryPath": "여성패션>니트", "isPrimary": "Y"}]}]},
        out, columns,
    )

    cats = read_form(out, columns)["products"][0]["카테고리"]
    assert len(cats) == 1
    assert cats[0]["categoryPath"] == "여성패션>니트"


def test_숨은_시트가_보존된다(prefilled_workbook, columns, tmp_path):
    out = tmp_path / "edited.xlsx"
    apply_changes(prefilled_workbook, {"변경": [{"상품키": "P-000001", "필드": {"brand": "NEW"}}]}, out, columns)

    wb = openpyxl.load_workbook(out)
    assert "_양식정보" in wb.sheetnames
    assert wb["_양식정보"].sheet_state == "veryHidden"
    assert read_form(out, columns)["exportId"] == "0198f3a1-1111-7000-8000-abcdefabcdef"


def test_신규_행은_모든_시트에_추가된다(prefilled_workbook, columns, tmp_path):
    out = tmp_path / "edited.xlsx"
    apply_changes(
        prefilled_workbook,
        {"신규": [{
            "상품키": "NEW-1",
            "필드": {"name": "새 니트", "basePrice": "29000"},
            "옵션": [{"optionKey": "G1", "optionName": "색상", "optionValueKey": "V9", "optionValueName": "검정"}],
            "조합": [{"조합": ["V9"], "variantCode": "SKU-K"}],
            "카테고리": [{"categoryPath": "여성패션>니트", "isPrimary": "Y"}],
        }]},
        out, columns,
    )

    products = {p["상품키"]: p for p in read_form(out, columns)["products"]}
    assert products["NEW-1"]["필드"]["name"] == "새 니트"
    assert products["NEW-1"]["조합"][0]["combination"] == "V9"
    assert len(products) == 2   # 기존 행이 살아 있다


def test_조합은_항상_정렬해_잇는다(prefilled_workbook, columns, tmp_path):
    """서버는 조합 중복을 문자열 원본으로 센다(bulk-draft.options.ts:228-234).
    정렬하지 않으면 V1+V2 와 V2+V1 이 갈려 같은 조합의 품목이 둘 생긴다."""
    out = tmp_path / "edited.xlsx"
    apply_changes(
        prefilled_workbook,
        {"신규": [{"상품키": "NEW-1", "필드": {"name": "x", "basePrice": "1"},
                   # check_workbook 의 ④ 참조 무결성(조합 키는 그 상품의 옵션값키 집합 안이어야
                   # 한다)을 만족시키려고 옵션을 같이 준다 — 이 테스트가 보는 것은 정렬이지 참조다.
                   "옵션": [
                       {"optionKey": "G1", "optionName": "색상", "optionValueKey": "V1", "optionValueName": "빨강"},
                       {"optionKey": "G1", "optionName": "색상", "optionValueKey": "V2", "optionValueName": "파랑"},
                   ],
                   "조합": [{"조합": ["V2", "V1"]}]}]},
        out, columns,
    )

    products = {p["상품키"]: p for p in read_form(out, columns)["products"]}
    assert products["NEW-1"]["조합"][0]["combination"] == "V1+V2"


def test_이미지_행이_추가된다(prefilled_workbook, columns, tmp_path):
    out = tmp_path / "edited.xlsx"
    apply_changes(
        prefilled_workbook,
        {"이미지": [{"imageKey": "IMG-10", "sourceValue": "NEW-1-main1.jpg"}]},
        out, columns,
    )

    assert read_form(out, columns)["이미지"] == [{"imageKey": "IMG-10", "sourceValue": "NEW-1-main1.jpg"}]


def test_없는_상품키를_변경하면_실패한다(prefilled_workbook, columns, tmp_path):
    with pytest.raises(ValueError, match="P-999999"):
        apply_changes(prefilled_workbook, {"변경": [{"상품키": "P-999999", "필드": {"brand": "x"}}]},
                      tmp_path / "edited.xlsx", columns)


def test_조합_대신_파생키_combination_을_직접_주면_거부된다(prefilled_workbook, columns, tmp_path):
    """read_form 의 출력이 바로 'combination' 키를 쓰므로, 읽고→고쳐서→되쓰는 흐름에서
    정렬을 건너뛴 문자열이 그대로 흘러들어오기 쉽다. 그 우회로를 문법 수준에서 막는다."""
    with pytest.raises(ValueError, match="combination"):
        apply_changes(
            prefilled_workbook,
            {"신규": [{"상품키": "NEW-1", "필드": {"name": "x", "basePrice": "1"},
                       "조합": [{"combination": "V2+V1", "variantCode": "SKU-X"}]}]},
            tmp_path / "edited.xlsx", columns,
        )


def test_서식만_있는_후행_행_뒤가_아니라_실제_마지막_데이터_행_바로_다음에_추가된다(columns, tmp_path):
    """openpyxl 은 서식만 있고 값은 없는 셀도 '사용된 범위'로 잡아 ws.max_row 를 부풀린다 —
    엑셀에서 열 전체를 선택해 색/테두리를 넣는 흔한 습관만으로 재현된다. 그 부풀려진 값을
    append 기준으로 쓰면 새 행이 빈 행 뭉치 뒤에 붙어 파일을 열었을 때 혼란스럽다."""
    defs = columns["sheets"]["상품"]
    key_col = {d["key"]: i + 1 for i, d in enumerate(defs)}  # 라벨이 아니라 열 순서로 추적

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = columns["sheetNames"]["products"]
    ws.append([d["label"] for d in defs])
    # 예약 형식(P-000001)을 쓰지 않는다: 이 워크북엔 _양식정보 시트가 없어 check_workbook 의
    # ⑥ 검사가 "숨은 시트를 잃은 흔적"으로 보고 걸어버린다 — 이 테스트가 보는 것은 그게 아니라
    # 서식만 있는 후행 행 뒤의 append 위치다.
    ws.append([{"rowKey": "SEED-001", "name": "티셔츠", "basePrice": "19000"}.get(d["key"], "")
                for d in defs])

    fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for r in range(3, 8):          # 값은 없이 서식만 있는 후행 행 5개
        for c in range(1, len(defs) + 1):
            ws.cell(row=r, column=c).fill = fill

    # check_workbook 이 무조건 들여다보는 나머지 시트들 — 헤더만 있어도 충분하다.
    names = columns["sheetNames"]
    for sheet_key, name_key in (("옵션", "options"), ("조합", "variants"),
                                 ("카테고리", "categories"), ("이미지", "images")):
        wb.create_sheet(names[name_key]).append([d["label"] for d in columns["sheets"][sheet_key]])

    src = tmp_path / "formatted.xlsx"
    wb.save(src)

    # 저장 직후 재확인: 이 테스트가 실제로 그 현상을 재현하고 있는지의 전제 확인.
    reloaded = openpyxl.load_workbook(src)
    assert reloaded[columns["sheetNames"]["products"]].max_row == 7

    out = tmp_path / "edited.xlsx"
    apply_changes(src, {"신규": [{"상품키": "NEW-1", "필드": {"name": "새 상품", "basePrice": "5000"}}]},
                  out, columns)

    result_ws = openpyxl.load_workbook(out)[columns["sheetNames"]["products"]]
    assert result_ws.cell(row=3, column=key_col["rowKey"]).value == "NEW-1"  # 빈 행 뭉치 뒤가 아니라 바로 다음 행

    products = {p["상품키"]: p for p in read_form(out, columns)["products"]}
    assert products["NEW-1"]["필드"]["name"] == "새 상품"
