"""테스트용 워크북 픽스처. 실제 양식과 같은 시트 이름·헤더를 쓴다."""
import json
import pathlib
import sys

import openpyxl
import pytest

SCRIPTS = pathlib.Path(__file__).resolve().parent.parent / "scripts"
sys.path.insert(0, str(SCRIPTS))


@pytest.fixture
def columns():
    return json.loads((SCRIPTS / "columns.json").read_text(encoding="utf-8"))


def _add_sheet(wb, name, columns, sheet_name, rows):
    ws = wb.create_sheet(name)
    defs = columns["sheets"][sheet_name]
    ws.append([c["label"] for c in defs])
    for row in rows:
        ws.append([row.get(c["key"], "") for c in defs])
    return ws


SHEET_KEYS = (
    ("상품", "products"), ("옵션", "options"), ("조합", "variants"),
    ("카테고리", "categories"), ("구매제약", "constraints"),
    ("이미지", "images"), ("카테고리 참조", "categoryReference"),
)


@pytest.fixture
def make_workbook(tmp_path, columns):
    """시트별 행 목록으로 워크북 하나를 만든다.

    `meta=False` 는 두 가지를 뜻할 수 있다 — 빈 양식(정상적으로 `_양식정보` 가 없다) 이거나
    숨은 시트를 잃은 프리필. 둘을 가르는 것은 상품키가 원본에 있었는지다.
    """
    def _make(name, rows_by_sheet=None, meta=True):
        rows_by_sheet = rows_by_sheet or {}
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for sheet_key, name_key in SHEET_KEYS:
            _add_sheet(wb, columns["sheetNames"][name_key], columns, sheet_key,
                       rows_by_sheet.get(sheet_key, []))
        if meta:
            ws = wb.create_sheet(columns["sheetNames"]["meta"])
            ws["A1"] = "exportId"
            ws["B1"] = "0198f3a1-1111-7000-8000-abcdefabcdef"
            ws.sheet_state = "veryHidden"
        path = tmp_path / name
        wb.save(path)
        return path
    return _make


@pytest.fixture
def prefilled_workbook(tmp_path, columns):
    """프리필 워크북 하나. 숨은 _양식정보 시트를 포함한다."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _add_sheet(wb, "상품", columns, "상품", [
        {"rowKey": "P-000001", "name": "티셔츠", "basePrice": "19000", "brand": "ACME"},
    ])
    _add_sheet(wb, "옵션", columns, "옵션", [
        {"rowKey": "P-000001", "optionKey": "G1", "optionName": "색상",
         "optionValueKey": "V1", "optionValueName": "빨강"},
        {"rowKey": "P-000001", "optionKey": "G1", "optionName": "색상",
         "optionValueKey": "V2", "optionValueName": "파랑"},
    ])
    _add_sheet(wb, "조합", columns, "조합", [
        {"rowKey": "P-000001", "combination": "V1", "variantCode": "SKU-R"},
        {"rowKey": "P-000001", "combination": "V2", "variantCode": "SKU-B"},
    ])
    _add_sheet(wb, "카테고리", columns, "카테고리", [
        {"rowKey": "P-000001", "categoryPath": "여성패션>티셔츠", "isPrimary": "Y"},
    ])
    _add_sheet(wb, "구매제약", columns, "구매제약", [])
    _add_sheet(wb, "이미지", columns, "이미지", [])
    _add_sheet(wb, "카테고리 참조", columns, "카테고리 참조", [
        {"categoryPath": "여성패션>티셔츠"},
        {"categoryPath": "여성패션>니트"},
    ])

    meta = wb.create_sheet("_양식정보")
    meta["A1"] = "exportId"
    meta["B1"] = "0198f3a1-1111-7000-8000-abcdefabcdef"
    meta.sheet_state = "veryHidden"

    path = tmp_path / "form.xlsx"
    wb.save(path)
    return path
