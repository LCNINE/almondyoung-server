#!/usr/bin/env python3
"""원본 워크북에 **변경분만** 적용해 새 파일을 만든다.

새 워크북을 만들지 않고 원본을 여는 것이 이 스크립트의 존재 이유다. 숨은 `_양식정보`
시트에 든 exportId 가 "이 워크북은 수정용"임을 말하는 유일한 표식이고, 그것을 잃은 파일을
올리면 프리필 행 전량이 신규 상품으로 재생성된다.

그리고 **전체 상태가 아니라 변경분을 받는다.** 적지 않은 키는 셀을 건드리지 않고, 비우려면
None 을 명시해야 한다 — 그래서 우연한 필드 비움이 문법 수준에서 불가능하다.
"""
import argparse
import json
import pathlib
import re

import openpyxl

from read_form import load_columns

META_SHEET = "_양식정보"
RESERVED_ROW_KEY_RE = re.compile(r"^P-\d{6}$")


def _header_map(ws, defs):
    """헤더 라벨 → 열 번호(1-based). 모르는 열은 담지 않는다."""
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    out = {}
    for i, cell in enumerate(header):
        label = "" if cell is None else str(cell).strip()
        for d in defs:
            if d["label"] == label:
                out[d["key"]] = i + 1
    return out


def _row_index_by_key(ws, header):
    col = header.get("rowKey")
    if col is None:
        return {}
    out = {}
    for r in range(2, ws.max_row + 1):
        value = ws.cell(row=r, column=col).value
        key = "" if value is None else str(value).strip()
        if key:
            out[key] = r
    return out


def _set_cells(ws, row_index, header, values):
    """None → 빈 문자열(명시적 비움). 키가 없으면 애초에 여기 오지 않는다(안 건드림)."""
    for key, value in values.items():
        col = header.get(key)
        if col is None:
            raise ValueError(f"'{ws.title}' 시트에 '{key}' 에 해당하는 열이 없습니다.")
        ws.cell(row=row_index, column=col).value = "" if value is None else str(value)


def _last_used_row(ws):
    """값이 실제로 있는 마지막 행 번호. 헤더뿐이면 1.

    `ws.max_row` 는 openpyxl 이 서식만 있고 값은 없는 셀도 "사용된 범위"에 넣어 계산한다 —
    엑셀에서 열 전체를 선택해 색이나 테두리를 넣는 흔한 습관만으로 부풀어, 왕복 후
    `ws.max_row` 가 실제 데이터 마지막 행보다 훨씬 커질 수 있다. 그 값을 그대로 append
    기준으로 쓰면 새 행이 빈 행 뭉치 뒤에 붙어 파일을 열었을 때 혼란스럽다.
    """
    for r in range(ws.max_row, 0, -1):
        if any(cell.value is not None for cell in ws[r]):
            return r
    return 1  # 헤더뿐인 빈 시트


def _append_row(ws, header, values, row_key=None):
    target = _last_used_row(ws) + 1
    if row_key is not None and "rowKey" in header:
        ws.cell(row=target, column=header["rowKey"]).value = row_key
    _set_cells(ws, target, header, values)
    return target


def _delete_rows_for_key(ws, header, row_key):
    col = header.get("rowKey")
    if col is None:
        return
    for r in range(ws.max_row, 1, -1):
        value = ws.cell(row=r, column=col).value
        if value is not None and str(value).strip() == row_key:
            ws.delete_rows(r)


def _normalize_variant(entry):
    """조합은 배열로 받아 **정렬해** '+' 로 잇는다.

    서버가 조합 중복을 문자열 원본으로 세기 때문이다(bulk-draft.options.ts:228-234).
    문자열을 그대로 받으면 A+B / B+A 가 갈려 같은 조합의 품목이 둘 만들어진다 —
    정렬 책임을 도구가 가져가면 그 사고가 날 자리가 없다.

    그래서 파생 출력 키인 'combination' 을 입력으로 직접 주는 것도 거부한다. `read_form`
    이 그 이름으로 결과를 내놓기 때문에, 읽고→고쳐서→되쓰는 흐름에서 그대로 흘러들어오기
    쉽다 — 그 경로를 열어두면 정렬 보장이 소용없어진다. 관대한 우회로를 새로 만들지 않는다.
    """
    if "combination" in entry:
        raise ValueError(
            "'조합' 항목에 파생 키 'combination' 을 직접 줄 수 없습니다. "
            "'조합': [옵션값키, ...] 형태의 배열로 주세요 — 정렬은 도구가 합니다."
        )
    out = dict(entry)
    combo = out.pop("조합", None)
    if combo is not None:
        if not isinstance(combo, list):
            raise ValueError("'조합' 은 옵션값키 배열이어야 합니다. 문자열을 직접 만들지 마세요.")
        out["combination"] = "+".join(sorted(str(c).strip() for c in combo if str(c).strip()))
    return out


CHILD_SHEETS = {"옵션": "options", "조합": "variants", "카테고리": "categories"}


def check_workbook(wb, columns, had_meta, original_keys, touched_keys):
    """양식 무결성만 본다. 비즈니스 규칙(가격·길이 상한·복합 가격규칙)은 서버 몫이다.

    검사를 얇게 유지하는 것이 설계다 — 서버 규칙을 미러하기 시작하면 그 미러가 조용히
    낡아 '스킬은 통과했는데 서버가 거부' 또는 더 나쁘게 '스킬이 유효한 작업을 막음' 이 된다.

    `original_keys`: apply_changes 가 변경을 적용하기 **전** `_row_index_by_key` 로 만든
    원본 상품키 집합. ⑥ 예약 상품키 검사에서만 쓴다 — "원본에 이미 있던 예약 키" 를
    가리려면 변경이 반영되기 전의 스냅샷이 필요하고, 이 함수에 넘어오는 `wb` 는 이미
    변경이 적용된 뒤이므로 그 자체로는 구분할 수 없다.

    `touched_keys`: 이번 변경 JSON 이 실제로 손댄 상품키(`변경` + `신규`).

    `(problems, warnings)` 를 돌려준다. ③④⑤ 는 **손댄 상품키에만 차단력을 쓴다** — 그
    밖의 행에서 발견한 이상은 경고로 실어 보내고 파일은 만든다. 서버가 같은 조건을
    행 단위로만 떨구고 세션은 진행시키기 때문이다(`bulk-upload.assembler.ts:81-88` 의
    고아 행 처리, `bulk-session.structure.ts:267-269` 의 대표 카테고리 RowError). 전
    워크북에 fail-closed 로 걸면 AI 가 건드리지도 않은 레거시 행 하나가 파일 생성을 전부
    막는데, 이 스크립트에는 행 삭제 기능이 없고 절대 규칙 1 이 새 워크북 생성을 금지하므로
    빠져나갈 길이 아예 없다. ①②⑥⑦⑧ 은 파일 수준 사실이거나(①②) 카탈로그 대량 중복
    방어선이거나(⑥) AI 자신의 실수(⑦⑧)라 전역 차단 그대로다.
    """
    problems = []
    warnings = []
    names, defs = columns["sheetNames"], columns["sheets"]

    def scoped(key, message):
        """이번에 손댄 상품이면 차단, 아니면 경고."""
        (problems if key in touched_keys else warnings).append(message)

    # ① 숨은 시트 보존
    if had_meta and META_SHEET not in wb.sheetnames:
        problems.append("양식 정보 시트(_양식정보)를 잃었습니다. 원본을 열어 셀만 고쳐야 합니다.")

    products_ws = wb[names["products"]]
    header = _header_map(products_ws, defs["상품"])

    # ② 필수 헤더
    for d in defs["상품"]:
        if d["required"] and d["key"] not in header:
            problems.append(f"'상품' 시트에 필수 열 '{d['label']}' 이 없습니다.")

    keys = []
    for r in range(2, products_ws.max_row + 1):
        value = products_ws.cell(row=r, column=header["rowKey"]).value if "rowKey" in header else None
        key = "" if value is None else str(value).strip()
        if key:
            keys.append(key)

    # ⑧ 상품키 중복
    for key in {k for k in keys if keys.count(k) > 1}:
        problems.append(f"상품키가 중복되었습니다: {key}")

    known = set(keys)

    # ③ 참조 무결성 + ④ 조합 키 + ⑤ 대표 카테고리
    values_by_product = {}
    for sheet_key, name_key in CHILD_SHEETS.items():
        ws = wb[names[name_key]]
        h = _header_map(ws, defs[sheet_key])
        if "rowKey" not in h:
            continue
        for r in range(2, ws.max_row + 1):
            raw = ws.cell(row=r, column=h["rowKey"]).value
            key = "" if raw is None else str(raw).strip()
            if not key:
                continue
            if key not in known:
                scoped(key, f"'{sheet_key}' 시트가 '상품' 시트에 없는 상품키를 참조합니다: {key}")
                continue
            if sheet_key == "옵션" and "optionValueKey" in h:
                v = ws.cell(row=r, column=h["optionValueKey"]).value
                if v:
                    values_by_product.setdefault(key, set()).add(str(v).strip())

    ws = wb[names["variants"]]
    h = _header_map(ws, defs["조합"])
    if "rowKey" in h and "combination" in h:
        for r in range(2, ws.max_row + 1):
            raw = ws.cell(row=r, column=h["rowKey"]).value
            key = "" if raw is None else str(raw).strip()
            combo = ws.cell(row=r, column=h["combination"]).value
            if not key or key not in known:
                continue   # ③ 이 이미 보고했다 — 같은 행을 두 번 세지 않는다
            if not combo:
                # 빈 조합은 "옵션 없는 상품의 단일 기본 품목"이라는 계약이 있어(서버
                # bulk-draft.options.ts:216) 그 자체로는 정상이다. 그러나 옵션 축이 있는
                # 상품에서 비면 어느 옵션값에도 안 묶인 품목이 되어 조용히 통과한다.
                if values_by_product.get(key):
                    scoped(key, f"'{key}' 는 옵션 행이 있는데 조합 행의 '조합' 이 비어 있습니다. "
                                "그 조합이 어떤 옵션값들의 묶음인지 지정해야 합니다.")
                continue
            for part in str(combo).split("+"):
                part = part.strip()
                if part and part not in values_by_product.get(key, set()):
                    scoped(key, f"'{key}' 의 조합이 옵션 시트에 없는 옵션값키를 참조합니다: {part}")

    ws = wb[names["categories"]]
    h = _header_map(ws, defs["카테고리"])
    if "rowKey" in h and "isPrimary" in h:
        primary = {}
        for r in range(2, ws.max_row + 1):
            raw = ws.cell(row=r, column=h["rowKey"]).value
            key = "" if raw is None else str(raw).strip()
            if not key or key not in known:
                continue   # ③ 이 이미 보고했다 — 같은 행을 두 번 세지 않는다
            flag = ws.cell(row=r, column=h["isPrimary"]).value
            primary.setdefault(key, 0)
            if flag is not None and str(flag).strip() == "Y":
                primary[key] += 1
        for key, count in primary.items():
            if count != 1:
                scoped(key, f"'{key}' 의 대표 카테고리는 정확히 1개여야 합니다 (현재 {count}개).")

    # ⑥ 예약 상품키 — 원본에 없던 예약 형식 키를 새로 만들면 서버가 거부한다.
    # `keys` 는 변경이 이미 적용된 워크북에서 읽은 것이라, 신규로 추가한 예약 형식 키도
    # 이미 섞여 들어와 있다. "원본에 이미 있던 키"인지는 `keys` 자체에서는 구별할 수
    # 없다 — 반드시 apply_changes 가 변경 적용 *전*에 넘겨준 original_keys 로 판별한다.
    #
    # 두 문구는 **원본에 있던 키인가**로 갈린다. had_meta 로 가르면 안 된다 — 빈 양식
    # (`빈 양식 다운로드`)은 정상적으로 `_양식정보` 시트가 없어서(form-export.workbook.ts:64-70)
    # had_meta 가 False 이고, 거기서 AI 가 예약 형식 키를 지으면 "양식을 다시 받아라"라는
    # 실행 불가능한 안내가 나간다(다시 받을 프리필이 없다).
    #
    # - 원본에 있던 예약 키 + 양식 정보 없음 → **숨은 _양식정보 시트를 잃은 프리필 워크북**.
    #   이 스킬 전체가 막으려는 시나리오다. "다른 상품키를 지어라"는 정반대 조치다 — 그대로
    #   따라 키를 바꿔 올리면 예약 형식이 아니게 되어 서버 가드(bulk-session.manager.ts)를
    #   비껴가고, 워크북이 "신규 전용 세션"으로 읽혀 프리필 행 전량이 대량 중복 생성된다.
    # - 원본에 있던 예약 키 + 양식 정보 있음 → 정상 프리필 행이다. 아무 문제 없다.
    # - 원본에 없던 예약 키 → AI 가 지은 것이다. 양식 정보 유무와 무관하게 다시 짓게 한다.
    for key in dict.fromkeys(keys):
        if not RESERVED_ROW_KEY_RE.match(key):
            continue
        if key not in original_keys:
            problems.append(f"'{key}' 는 시스템 예약 상품키 형식입니다. 다른 상품키를 지어 주세요.")
        elif not had_meta:
            problems.append(
                f"'{key}' 는 시스템이 발급한 상품키인데 이 파일에는 양식 정보(_양식정보)가 없습니다. "
                "상품키를 바꾸지 말고, 상품 목록에서 양식을 다시 받아 작성해 주세요."
            )

    # ⑦ 이미지 URL
    ws = wb[names["images"]]
    h = _header_map(ws, defs["이미지"])
    if "sourceValue" in h:
        for r in range(2, ws.max_row + 1):
            value = ws.cell(row=r, column=h["sourceValue"]).value
            text = "" if value is None else str(value).strip()
            if text.startswith("http://") or text.startswith("https://"):
                problems.append(f"이미지 원본에 URL 은 쓸 수 없습니다: {text}")

    return problems, warnings


def apply_changes(src_xlsx, changes, out_xlsx, columns):
    wb = openpyxl.load_workbook(src_xlsx)
    had_meta = META_SHEET in wb.sheetnames

    names = columns["sheetNames"]
    defs = columns["sheets"]

    products_ws = wb[names["products"]]
    products_header = _header_map(products_ws, defs["상품"])
    row_index = _row_index_by_key(products_ws, products_header)
    # 변경을 적용하기 전의 스냅샷 — row_index 는 아래에서 신규 상품키를 추가하며 그대로
    # 오염되므로, "원본에 있던 키"를 나중에 다시 물어볼 방법이 이것 말고 없다.
    original_keys = set(row_index.keys())

    report = {"변경": 0, "신규": 0, "이미지": 0, "경고": []}
    touched_keys = set()

    for change in changes.get("변경", []):
        row_key = change["상품키"]
        touched_keys.add(row_key)
        if row_key not in row_index:
            raise ValueError(f"'{row_key}' 상품키가 원본 양식에 없습니다. 수정 대상이 맞는지 확인하세요.")
        _set_cells(products_ws, row_index[row_key], products_header, change.get("필드", {}))

        for sheet_key, name_key in CHILD_SHEETS.items():
            if sheet_key not in change:
                continue   # 안 주면 안 건드린다
            ws = wb[names[name_key]]
            header = _header_map(ws, defs[sheet_key])
            _delete_rows_for_key(ws, header, row_key)
            for entry in change[sheet_key]:
                _append_row(ws, header, _normalize_variant(entry), row_key)
        report["변경"] += 1

    for product in changes.get("신규", []):
        row_key = product["상품키"]
        touched_keys.add(row_key)
        if row_key in row_index:
            raise ValueError(f"'{row_key}' 상품키가 이미 양식에 있습니다. 신규 행의 상품키는 유일해야 합니다.")
        # products_ws.max_row 는 서식만 있는 후행 행 탓에 부풀 수 있으므로(_last_used_row 참고),
        # 방금 쓴 실제 행 번호는 반환값에서 받는다 — max_row 를 다시 읽지 않는다.
        row_index[row_key] = _append_row(products_ws, products_header, product.get("필드", {}), row_key)

        for sheet_key, name_key in CHILD_SHEETS.items():
            for entry in product.get(sheet_key, []):
                ws = wb[names[name_key]]
                _append_row(ws, _header_map(ws, defs[sheet_key]), _normalize_variant(entry), row_key)
        report["신규"] += 1

    if "이미지" in changes:
        ws = wb[names["images"]]
        header = _header_map(ws, defs["이미지"])
        for entry in changes["이미지"]:
            _append_row(ws, header, entry)
            report["이미지"] += 1

    problems, warnings = check_workbook(wb, columns, had_meta, original_keys, touched_keys)
    if problems:
        # **파일을 만들지 않는다.** 반쯤 맞는 워크북을 손에 쥐면 사람이 그걸 올려버린다.
        raise ValueError("양식 검사에서 문제를 찾았습니다:\n- " + "\n- ".join(problems))

    wb.save(out_xlsx)
    report["경고"] = warnings
    return report


def main():
    parser = argparse.ArgumentParser(description="원본 양식에 변경분을 적용한다")
    parser.add_argument("src")
    parser.add_argument("changes", help="변경 JSON 파일 경로")
    parser.add_argument("out")
    parser.add_argument("--columns", default=str(pathlib.Path(__file__).parent / "columns.json"))
    args = parser.parse_args()

    changes = json.loads(pathlib.Path(args.changes).read_text(encoding="utf-8"))
    report = apply_changes(args.src, changes, args.out, load_columns(args.columns))
    print(f"변경 {report['변경']}건 · 신규 {report['신규']}건 · 이미지 {report['이미지']}건 → {args.out}")
    if report["경고"]:
        print(
            f"\n경고 {len(report['경고'])}건 — 이번에 손대지 않은 상품의 행에서 발견했습니다. "
            "파일은 만들었고, 서버도 이런 행만 오류로 떨구고 나머지는 진행합니다. "
            "그대로 두려면 두고, 고치려면 원본 워크북에서 그 행을 손봐야 합니다:"
        )
        for warning in report["경고"]:
            print(f"- {warning}")


if __name__ == "__main__":
    main()
