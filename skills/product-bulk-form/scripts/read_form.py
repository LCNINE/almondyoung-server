#!/usr/bin/env python3
"""양식 워크북을 상품 단위 JSON 으로 읽는다.

이 스크립트의 본질은 **조인**이다. 워크북은 상품키로 이어진 6개 시트지만 AI 는 상품 하나가
객체 하나인 형태로 다뤄야 시트 간 참조를 손으로 따라다니지 않는다.

라벨은 하드코딩하지 않는다 — 옆의 columns.json 에서 읽는다. 그 파일은 서버 코드
(form-export.sheets.ts)에서 생성되므로 열이 바뀌어도 갈리지 않는다.
"""
import argparse
import json
import pathlib

import openpyxl

META_SHEET = "_양식정보"
META_CELL = "B1"


def load_columns(path):
    return json.loads(pathlib.Path(path).read_text(encoding="utf-8"))


def _read_sheet(wb, sheet_name, defs):
    """헤더 '이름'으로 열을 찾는다 — 열 순서는 자유이고 모르는 열은 무시한다."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = ["" if c is None else str(c).strip() for c in rows[0]]
    key_by_index = {}
    for i, label in enumerate(header):
        for d in defs:
            if d["label"] == label:
                key_by_index[i] = d["key"]

    out = []
    for raw in rows[1:]:
        cells = {}
        has_value = False
        for i, value in enumerate(raw):
            key = key_by_index.get(i)
            if key is None:
                continue
            text = "" if value is None else str(value).strip()
            cells[key] = text
            if text:
                has_value = True
        if has_value:
            out.append(cells)
    return out


def read_form(xlsx_path, columns):
    wb = openpyxl.load_workbook(xlsx_path)
    names = columns["sheetNames"]
    sheets = columns["sheets"]

    export_id = None
    if META_SHEET in wb.sheetnames:
        value = wb[META_SHEET][META_CELL].value
        export_id = str(value).strip() if value else None

    products = _read_sheet(wb, names["products"], sheets["상품"])
    children = {
        "옵션": _read_sheet(wb, names["options"], sheets["옵션"]),
        "조합": _read_sheet(wb, names["variants"], sheets["조합"]),
        "카테고리": _read_sheet(wb, names["categories"], sheets["카테고리"]),
        "구매제약": _read_sheet(wb, names["constraints"], sheets["구매제약"]),
    }

    by_key = {}
    out_products = []
    for cells in products:
        row_key = cells.get("rowKey", "")
        product = {
            "상품키": row_key,
            # exportId 가 없으면 프리필 자체가 존재하지 않는다(빈 양식).
            "출처": "프리필" if export_id else "신규",
            "필드": {k: v for k, v in cells.items() if k != "rowKey"},
            "옵션": [], "조합": [], "카테고리": [], "구매제약": None,
        }
        by_key[row_key] = product
        out_products.append(product)

    # 상품 시트에 없는 상품키를 가리키는 자식 행은 조용히 버리지 않는다. 버리면 AI 는 읽기
    # 결과에서 그 행을 볼 수 없는데 서버는 업로드 후 그 행을 오류로 떨구고(assembler 의
    # `"상품" 시트에 없는 상품키를 참조했습니다`), `write_form.py` 도 같은 행을 경고한다 —
    # 원인이 보이지 않는 막다른 길이 된다. `products` 구조는 그대로 두고 최상위로만 드러낸다.
    orphans = []
    for sheet_key in ("옵션", "조합", "카테고리", "구매제약"):
        for cells in children[sheet_key]:
            target = by_key.get(cells.get("rowKey", ""))
            if target is None:
                orphans.append({"시트": sheet_key, **cells})   # rowKey 를 남긴다 — 그게 단서다
            elif sheet_key == "구매제약":
                target["구매제약"] = {k: v for k, v in cells.items() if k != "rowKey"}
            else:
                target[sheet_key].append({k: v for k, v in cells.items() if k != "rowKey"})

    return {
        "exportId": export_id,
        "products": out_products,
        "고아행": orphans,
        "이미지": _read_sheet(wb, names["images"], sheets["이미지"]),
        "카테고리참조": [
            c.get("categoryPath", "")
            for c in _read_sheet(wb, names["categoryReference"], sheets["카테고리 참조"])
        ],
    }


def main():
    parser = argparse.ArgumentParser(description="양식 워크북을 JSON 으로 읽는다")
    parser.add_argument("xlsx")
    parser.add_argument("--columns", default=str(pathlib.Path(__file__).parent / "columns.json"))
    parser.add_argument("--out")
    args = parser.parse_args()

    data = read_form(args.xlsx, load_columns(args.columns))
    text = json.dumps(data, ensure_ascii=False, indent=2)
    if args.out:
        pathlib.Path(args.out).write_text(text, encoding="utf-8")
        summary = f"wrote {args.out}: 상품 {len(data['products'])}건"
        if data["고아행"]:
            summary += f" · 고아행 {len(data['고아행'])}건(상품 시트에 없는 상품키를 가리킴)"
        print(summary)
    else:
        print(text)


if __name__ == "__main__":
    main()
