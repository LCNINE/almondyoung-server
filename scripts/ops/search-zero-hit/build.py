# -*- coding: utf-8 -*-
"""엑셀 리포트 생성 — 3단계.

collect.py 의 수집 결과와 naver.js 의 시중 판매 확인 결과를 합쳐 판정하고,
소싱 후보 / 노출 결함 / 타 업종으로 갈라 엑셀로 떨어뜨린다.

  python3 build.py --out ~/Downloads/아몬드영_검색0건_미취급_키워드.xlsx

판정 사전은 verdicts.json 이다. 오탐을 발견하면 그 파일에 키워드를 추가하면
다음 실행부터 반영된다 — 코드는 건드리지 않는다.
"""
import argparse, json, os, re
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DIR = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(DIR, "out")

# 우리 업계 = 속눈썹 / 반영구 / 타투 / 네일 / 왁싱 / 헤어 전문 / 피부관리 재료
TRADE = re.compile(
    r"속눈썹|속눈섭|래쉬|lash|아이래쉬|눈썹|글루|롯드|롯트|펌제|펌지|연장모|가모|네일|젤네일|폴리쉬|큐티클|"
    r"타투|문신|반영구|엠보|색소|피그먼트|니들|카트리지|왁싱|제모|왁스워머|시술|핀셋|트위저|글루판|전처리|"
    r"프라이머|리무버|아이패치|마취|마스카라|펌|미용실|이발기|바리깡|염색약|헤나|헨나|흑채|포마드|파마|두피|"
    r"에스테틱|피부관리|모델링팩|고무팩|각질|필링", re.I)
GENERIC = re.compile(
    r"로봇청소기|선풍기|냉장고|자동차|굴삭기|피규어|인형|와인|쌀|주먹밥|김밥|아이스크림|성경|프라모델|책상|"
    r"침대|매트리스|현수막|배드민턴|씨앗|모종|굴착|나사|볼트|복스|전자책|학습지|콜라텍|건강식품", re.I)


def classify(kw, items, verdicts):
    if kw in verdicts["trade"]:
        return "소싱 후보", verdicts["trade"][kw], ""
    if not items:
        return "시중 확인 불가", "", ""
    if kw in verdicts["other"]:
        return "타 업종", items[0]["n"], items[0]["c"]
    beauty = [i for i in items if i["c"].startswith("화장품/미용")]
    trade = [i for i in items if TRADE.search(i["n"]) and not GENERIC.search(i["n"])]
    both = [i for i in trade if i in beauty]
    if both:
        return "소싱 후보", both[0]["n"], both[0]["c"]
    if trade:
        return "소싱 후보", trade[0]["n"], trade[0]["c"]
    if beauty:
        return "업계 인접", beauty[0]["n"], beauty[0]["c"]
    return "타 업종", items[0]["n"], items[0]["c"]


HEAD = PatternFill("solid", fgColor="FFF2E5")


def sheet(wb, title, note, header, rows, widths):
    ws = wb.create_sheet(title)
    ws["A1"] = note
    ws["A1"].font = Font(bold=True, color="996600")
    ws.append([])
    ws.append(header)
    for c in ws[3]:
        c.font, c.fill = Font(bold=True), HEAD
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"
    for row in ws.iter_rows(min_row=4):
        for c in row:
            c.alignment = Alignment(vertical="center")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=os.path.expanduser("~/Downloads/아몬드영_검색0건_미취급_키워드.xlsx"))
    a = ap.parse_args()

    col = json.load(open(f"{OUT}/collected.json"))
    nb = json.load(open(f"{OUT}/naver.json"))
    verdicts = json.load(open(f"{DIR}/verdicts.json"))
    verdicts["other"] = set(verdicts["other"])

    period = f'{col["since"]} ~ {col["until"] if col["until"] != "2100-01-01" else "오늘"}'
    buckets = {}
    for r in col["missing"]:
        v, name, cat = classify(r["kw"], (nb.get(r["kw"]) or {}).get("items") or [], verdicts)
        buckets.setdefault(v, []).append([r["kw"], r["cnt"], r["first"], r["last"], name, cat])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    HD = ["키워드", "0건 검색수", "첫 검색", "마지막 검색", "시중 상품 예시", "시중 카테고리"]
    W = [24, 11, 12, 12, 66, 34]

    sheet(wb, "소싱 후보",
          f"★ {period}, 우리 사이트 검색 결과가 0건이면서 네이버쇼핑에는 같은 이름의 업계 상품이 실제로 판매 중인 키워드입니다.",
          HD, buckets.get("소싱 후보", []), W)
    sheet(wb, "업계 인접",
          "미용 카테고리 상품이 나오지만 일반 소비자용 화장품에 가까운 키워드입니다. 취급 여부는 판단이 필요합니다.",
          HD, buckets.get("업계 인접", []), W)
    sheet(wb, "타 업종",
          "검색은 되지만 우리 업계와 무관한 상품이 나오는 키워드입니다. 조치 불필요.",
          HD, buckets.get("타 업종", []), W)
    sheet(wb, "시중 확인 불가",
          "네이버쇼핑에서 상품이 확인되지 않은 키워드입니다. 네이버가 검색을 막은 키워드(의료 관련 등)도 여기 섞이므로, 업계 용어로 보이면 직접 확인해 verdicts.json 에 넣어주세요.",
          HD, buckets.get("시중 확인 불가", []), W)

    sheet(wb, "노출 결함 0건",
          "★ 미취급이 아닙니다. 상품은 있는데 검색에 안 잡히는 키워드입니다. 대부분 멤버십 전용이라 비회원에게 0건으로 보이는 경우 — 소싱이 아니라 노출 정책 검토 대상입니다.",
          ["키워드", "0건 검색수", "보유 상품 수", "대표 상품", "비고"],
          [[r["kw"], r["cnt"], r["idx"], (r["idx_items"][0]["n"] if r["idx_items"] else ""),
            ("멤버십 전용" if any(i["mem"] for i in r["idx_items"]) else "")]
           for r in sorted(col["defect"], key=lambda x: -x["cnt"])],
          [22, 11, 12, 60, 16])

    sheet(wb, "이미 해결됨",
          "과거엔 0건이었지만 지금은 상품이 검색되는 키워드입니다(입고 또는 오타 교정 개선). 조치 불필요 — 참고용.",
          ["키워드", "과거 0건 검색수", "현재 검색 결과", "마지막 0건 검색"],
          [[r["kw"], r["cnt"], r["now"], r["last"]] for r in sorted(col["solved"], key=lambda x: -x["cnt"])],
          [22, 14, 13, 14])

    wb.save(a.out)
    print("저장:", a.out)
    for k, v in sorted(buckets.items(), key=lambda x: -len(x[1])):
        print(f"  {k} {len(v)}")
    print(f"  노출 결함 {len(col['defect'])} | 이미 해결됨 {len(col['solved'])}")


if __name__ == "__main__":
    main()
